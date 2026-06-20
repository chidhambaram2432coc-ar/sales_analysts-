"""
Retail Sales & Revenue Performance Dashboard
Flask Backend — app.py
"""

import sqlite3
import os
import io
from datetime import date
from flask import Flask, jsonify, render_template, request, send_file
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
DB_PATH = "retail_analytics.db"


# ---------------------------------------------------------------------------
# Database helpers
# ---------------------------------------------------------------------------

def get_db():
    """Open a DB connection with named-column row factory."""
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    """Create the sales_transactions table and seed baseline rows if empty."""
    conn = get_db()
    cur = conn.cursor()

    cur.execute("""
        CREATE TABLE IF NOT EXISTS sales_transactions (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            order_date   TEXT    NOT NULL,
            product_name TEXT    NOT NULL,
            category     TEXT    NOT NULL,
            region       TEXT    NOT NULL,
            sales_amount REAL    NOT NULL,
            quantity     INTEGER NOT NULL,
            profit       REAL    NOT NULL,
            status       TEXT    NOT NULL
        )
    """)

    cur.execute("SELECT COUNT(*) AS cnt FROM sales_transactions")
    if cur.fetchone()["cnt"] == 0:
        baseline = [
            ("2024-01-05", "MacBook Pro 14",      "Electronics", "North",  2499.00, 1, 374.85, "Delivered"),
            ("2024-01-12", "Levi's 501 Jeans",    "Clothing",    "South",   89.50, 3,  13.43, "Shipped"),
            ("2024-02-03", "Dyson V15 Vacuum",    "Appliances",  "East",   699.99, 2, 104.99, "Delivered"),
            ("2024-02-18", "Organic Coffee Beans","Groceries",   "West",    45.00,10,   6.75, "Pending"),
            ("2024-03-07", "Sony WH-1000XM5",     "Electronics", "South",  349.00, 4,  52.35, "Delivered"),
            ("2024-03-14", "IKEA KALLAX Shelf",   "Furniture",   "North",  189.00, 2,  28.35, "Shipped"),
            ("2024-04-01", "Nike Air Max 270",    "Clothing",    "East",   150.00, 5,  22.50, "Delivered"),
            ("2024-04-20", "Samsung 4K Monitor",  "Electronics", "West",   599.00, 1,  89.85, "Pending"),
        ]
        cur.executemany(
            """INSERT INTO sales_transactions
               (order_date, product_name, category, region, sales_amount, quantity, profit, status)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
            baseline,
        )
        conn.commit()
        print(f"  ✓ Seeded {len(baseline)} baseline rows.")
    else:
        print("  ✓ Database already populated — skipping seed.")

    conn.close()


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------

@app.route("/")
def index():
    """Render the main dashboard page."""
    return render_template("index.html")


@app.route("/api/submit_transaction", methods=["POST"])
def submit_transaction():
    """
    Accept a new sales transaction via JSON POST.

    Required JSON fields:
        product_name, category, region, sales_amount, quantity, status

    The backend automatically:
        • Injects today's date as order_date
        • Computes profit = sales_amount × 0.15
    """
    data = request.get_json(force=True)

    required = ["product_name", "category", "region", "sales_amount", "quantity", "status"]
    missing = [f for f in required if f not in data]
    if missing:
        return jsonify({"error": f"Missing fields: {', '.join(missing)}"}), 400

    try:
        sales_amount = float(data["sales_amount"])
        quantity     = int(data["quantity"])
    except (ValueError, TypeError):
        return jsonify({"error": "sales_amount must be numeric; quantity must be integer."}), 400

    # Algorithmic profit rule: exactly 15 % of sales_amount
    profit     = round(sales_amount * 0.15, 2)
    order_date = date.today().isoformat()

    conn = get_db()
    cur  = conn.cursor()
    cur.execute(
        """INSERT INTO sales_transactions
           (order_date, product_name, category, region, sales_amount, quantity, profit, status)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
        (order_date, data["product_name"], data["category"],
         data["region"], sales_amount, quantity, profit, data["status"]),
    )
    conn.commit()
    new_id = cur.lastrowid
    conn.close()

    return jsonify({"success": True, "id": new_id, "profit_computed": profit}), 201


@app.route("/api/dashboard_data", methods=["GET"])
def dashboard_data():
    """
    Return real-time KPIs and chart datasets computed via Pandas aggregation.

    Response shape:
    {
      "kpis": { total_revenue, net_profit, total_orders },
      "regional_sales": { "North": x, "South": y, ... },
      "fulfillment_status": { "Delivered": n, "Shipped": n, "Pending": n }
    }
    """
    conn = get_db()
    df   = pd.read_sql_query("SELECT * FROM sales_transactions", conn)
    conn.close()

    if df.empty:
        return jsonify({
            "kpis": {"total_revenue": 0, "net_profit": 0, "total_orders": 0},
            "regional_sales": {},
            "fulfillment_status": {},
        })

    # --- KPIs ---
    total_revenue = round(float(df["sales_amount"].sum()), 2)
    net_profit    = round(float(df["profit"].sum()), 2)
    total_orders  = int(len(df))

    # --- Regional Sales Distribution ---
    regional_sales = (
        df.groupby("region")["sales_amount"]
          .sum()
          .round(2)
          .to_dict()
    )

    # --- Order Fulfillment Status ---
    fulfillment_status = (
        df["status"]
          .value_counts()
          .to_dict()
    )

    return jsonify({
        "kpis": {
            "total_revenue": total_revenue,
            "net_profit":    net_profit,
            "total_orders":  total_orders,
        },
        "regional_sales":    regional_sales,
        "fulfillment_status": fulfillment_status,
    })


# ---------------------------------------------------------------------------
# Export Routes
# ---------------------------------------------------------------------------

@app.route("/api/export/csv", methods=["GET"])
def export_csv():
    """Export all sales_transactions as a CSV file download."""
    conn = get_db()
    df   = pd.read_sql_query("SELECT * FROM sales_transactions ORDER BY id DESC", conn)
    conn.close()

    buf = io.StringIO()
    df.to_csv(buf, index=False)
    buf.seek(0)

    return send_file(
        io.BytesIO(buf.getvalue().encode("utf-8")),
        mimetype="text/csv",
        as_attachment=True,
        download_name=f"retail_sales_{date.today().isoformat()}.csv",
    )


@app.route("/api/export/excel", methods=["GET"])
def export_excel():
    """Export all sales_transactions as a formatted Excel (.xlsx) file download."""
    conn = get_db()
    df   = pd.read_sql_query("SELECT * FROM sales_transactions ORDER BY id DESC", conn)
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Transactions"

    # ── Styles ──
    header_font    = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill    = PatternFill("solid", start_color="1E293B")
    center_align   = Alignment(horizontal="center", vertical="center")
    currency_fmt   = '#,##0.00'
    thin_border    = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )

    # ── Header row ──
    columns = list(df.columns)
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name.upper().replace("_", " "))
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center_align

    # ── Data rows ──
    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
            # Currency format for money columns
            col_name = columns[col_idx - 1]
            if col_name in ("sales_amount", "profit"):
                cell.number_format = currency_fmt

    # ── KPI Summary sheet ──
    ws_kpi = wb.create_sheet("KPI Summary")
    kpi_header_fill = PatternFill("solid", start_color="4F46E5")

    kpis = [
        ("Total Sales Revenue", f"=SUM('Sales Transactions'!F:F)"),
        ("Net Yield Profit",    f"=SUM('Sales Transactions'!H:H)"),
        ("Total Orders",        f"=COUNTA('Sales Transactions'!A:A)-1"),
    ]

    ws_kpi["A1"] = "KPI"
    ws_kpi["B1"] = "VALUE"
    for cell in [ws_kpi["A1"], ws_kpi["B1"]]:
        cell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        cell.fill      = kpi_header_fill
        cell.alignment = center_align

    for i, (label, formula) in enumerate(kpis, 2):
        ws_kpi.cell(row=i, column=1, value=label).font = Font(name="Arial", bold=True)
        val_cell = ws_kpi.cell(row=i, column=2, value=formula)
        if i < 4:
            val_cell.number_format = currency_fmt

    # ── Column widths ──
    col_widths = {"id":4,"order_date":14,"product_name":26,"category":14,
                  "region":10,"sales_amount":16,"quantity":10,"profit":14,"status":12}
    for col_idx, col_name in enumerate(columns, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_name, 14)

    ws_kpi.column_dimensions["A"].width = 24
    ws_kpi.column_dimensions["B"].width = 18

    # ── Freeze header row ──
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"retail_sales_{date.today().isoformat()}.xlsx",
    )


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    print("Initialising database…")
    init_db()
    print("Starting Retail Sales Analytics → http://127.0.0.1:5000\n")
    app.run(debug=True)
